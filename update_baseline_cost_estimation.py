#!/usr/bin/env python3
"""Upsert baseline pull-up and pushdown q-error means into the shared results workbook."""

from __future__ import annotations

import argparse
import os
from typing import Dict, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from summarize_repeated_runs import aggregate, read_workbooks


METRICS = ("q50", "q95", "q99")
WORKLOAD_PREFIXES = (("pullup", "pull_up"), ("pushdown", "push_down"))
HEADERS = [
    "test_db",
    "cardinality_type",
    *(f"{prefix}_{metric}_mean" for _, prefix in WORKLOAD_PREFIXES for metric in METRICS),
    "time_stamp",
    "epochs",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("summary_xlsx", nargs="+", help="Per-run summary XLSX files")
    parser.add_argument("--requested-runs", type=int, required=True)
    parser.add_argument("--test-db", required=True)
    parser.add_argument("--cardinality-type", required=True)
    parser.add_argument("--time-stamp", required=True)
    parser.add_argument("--epochs", type=int, required=True)
    parser.add_argument("--output-xlsx", required=True)
    return parser.parse_args()


def workload_kind(workload: object) -> Optional[str]:
    name = str(workload or "").casefold()
    for kind, _ in WORKLOAD_PREFIXES:
        if kind in name:
            return kind
    return None


def build_values(args: argparse.Namespace, summary: Dict[str, object]) -> Dict[str, object]:
    values = {
        "test_db": args.test_db,
        "cardinality_type": args.cardinality_type,
        "time_stamp": args.time_stamp,
        "epochs": args.epochs,
    }
    workload_metrics = {}
    for workload, metrics in summary["workloads"].items():
        kind = workload_kind(workload)
        if kind is not None:
            workload_metrics[kind] = metrics
    for kind, prefix in WORKLOAD_PREFIXES:
        metrics = workload_metrics.get(kind, {})
        for metric in METRICS:
            values[f"{prefix}_{metric}_mean"] = metrics.get(metric, (None, None, 0))[0]
    return values


def read_existing_results(worksheet):
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    header_index = {str(header): index for index, header in enumerate(headers) if header is not None}
    results = []
    for row in rows:
        result = {
            header: row[header_index[header]] if header in header_index else None
            for header in HEADERS
        }
        if result["test_db"] not in (None, ""):
            results.append(result)
    return results


def upsert_result(path: str, values: Dict[str, object]) -> None:
    output_dir = os.path.dirname(path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    if os.path.exists(path):
        workbook = load_workbook(path)
        worksheet = workbook["Baseline"] if "Baseline" in workbook.sheetnames else workbook.active
        worksheet.title = "Baseline"
        if "BaselineWorkloads" in workbook.sheetnames:
            workbook.remove(workbook["BaselineWorkloads"])
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Baseline"

    results = read_existing_results(worksheet)
    target_index = None
    for index, result in enumerate(results):
        same_database = str(result["test_db"] or "").casefold() == str(values["test_db"]).casefold()
        existing_cardinality = str(result["cardinality_type"] or "")
        same_cardinality = existing_cardinality.casefold() == str(values["cardinality_type"]).casefold()
        legacy_row = existing_cardinality == ""
        if same_database and (same_cardinality or legacy_row):
            target_index = index
            break
    if target_index is None:
        results.append(values)
    else:
        results[target_index] = values

    worksheet.delete_rows(1, worksheet.max_row)
    worksheet.append(HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for result in results:
        worksheet.append(["" if result.get(header) is None else result.get(header) for header in HEADERS])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column, header in enumerate(HEADERS, start=1):
        width = max(len(header) + 2, 14 if header != "time_stamp" else 18)
        worksheet.column_dimensions[worksheet.cell(row=1, column=column).column_letter].width = width
    workbook.save(path)
    workbook.close()


def main() -> int:
    args = parse_args()
    rows, run_details = read_workbooks(args.summary_xlsx)
    summary = aggregate(rows, args.requested_runs, run_details)
    if summary["successful_runs"] == 0:
        print("Baseline result not updated: no successful runs were found.")
        return 1

    values = build_values(args, summary)
    missing_metrics = [
        header for header in HEADERS
        if (header.startswith("pull_up_") or header.startswith("push_down_"))
        and values.get(header) is None
    ]
    if missing_metrics:
        print("Baseline result not updated: missing workload mean(s): " + ", ".join(missing_metrics))
        return 1
    upsert_result(args.output_xlsx, values)
    print(f"Baseline result updated: {args.output_xlsx} ({args.test_db}, {args.cardinality_type})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
