#!/usr/bin/env python3
"""Upsert augmented q-error means and baseline differences into a results workbook."""

from __future__ import annotations

import argparse
import math
import os
from pathlib import Path
from typing import Dict, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from summarize_repeated_runs import aggregate, read_workbooks


METRICS = ("q50", "q95", "q99")
WORKLOAD_PREFIXES = (("pullup", "pull_up"), ("pushdown", "push_down"))
CONFIG_HEADERS = (
    "test-augment",
    "augment-pooling",
    "augment-refinement",
    "augment-coarse-layers",
    "augment-include-inv",
    "augment-refine-ret",
    "lambda-struct",
)
HEADERS = [
    "test_db",
    "cardinality_type",
    *(f"{prefix}_{metric}_mean" for _, prefix in WORKLOAD_PREFIXES for metric in METRICS),
    *(f"diff_{prefix}_{metric}_mean" for _, prefix in WORKLOAD_PREFIXES for metric in METRICS),
    "time_stamp",
    *CONFIG_HEADERS,
    "epochs",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("summary_xlsx", nargs="+", help="Per-run augmented summary XLSX files")
    parser.add_argument("--requested-runs", type=int, required=True)
    parser.add_argument("--test-db", required=True)
    parser.add_argument("--cardinality-type", required=True)
    parser.add_argument("--time-stamp", required=True)
    parser.add_argument("--epochs", type=int, required=True)
    parser.add_argument("--baseline-xlsx", required=True)
    parser.add_argument("--output-xlsx", required=True)
    parser.add_argument("--test-augment", required=True)
    parser.add_argument("--augment-pooling", required=True)
    parser.add_argument("--augment-refinement", required=True)
    parser.add_argument("--augment-coarse-layers", type=int, required=True)
    parser.add_argument("--augment-include-inv", required=True)
    parser.add_argument("--augment-refine-ret", required=True)
    parser.add_argument("--lambda-struct", type=float, required=True)
    return parser.parse_args()


def as_float(value: object) -> Optional[float]:
    if value is None or str(value).strip() == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def read_baseline(path: Path, test_db: str, cardinality_type: str) -> Optional[Dict[str, Dict[str, float]]]:
    if not path.exists():
        return None

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Baseline"] if "Baseline" in workbook.sheetnames else workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        workbook.close()
        return None

    header_index = {str(header): index for index, header in enumerate(headers) if header is not None}
    required = {
        "test_db",
        "cardinality_type",
        *(f"{prefix}_{metric}_mean" for _, prefix in WORKLOAD_PREFIXES for metric in METRICS),
    }
    if not required.issubset(header_index):
        workbook.close()
        return None

    result = None
    for row in rows:
        database = str(row[header_index["test_db"]] or "")
        cardinality = str(row[header_index["cardinality_type"]] or "")
        if database.casefold() != test_db.casefold() or cardinality.casefold() != cardinality_type.casefold():
            continue
        values = {
            kind: {
                metric: as_float(row[header_index[f"{prefix}_{metric}_mean"]])
                for metric in METRICS
            }
            for kind, prefix in WORKLOAD_PREFIXES
        }
        if all(value is not None for metrics in values.values() for value in metrics.values()):
            result = values
        break

    workbook.close()
    return result


def workload_kind(workload: object) -> Optional[str]:
    name = str(workload or "").casefold()
    for kind, _ in WORKLOAD_PREFIXES:
        if kind in name:
            return kind
    return None


def build_values(
        args: argparse.Namespace,
        summary: Dict[str, object],
        baseline: Dict[str, Dict[str, float]]) -> Dict[str, object]:
    augmented = {}
    for workload, metrics in summary["workloads"].items():
        kind = workload_kind(workload)
        if kind is None:
            continue
        augmented[kind] = {
            metric: as_float(metrics.get(metric, (None, None, 0))[0])
            for metric in METRICS
        }
    values = {
        "test_db": args.test_db,
        "cardinality_type": args.cardinality_type,
        "time_stamp": args.time_stamp,
        "epochs": args.epochs,
        "test-augment": args.test_augment,
        "augment-pooling": args.augment_pooling,
        "augment-refinement": args.augment_refinement,
        "augment-coarse-layers": args.augment_coarse_layers,
        "augment-include-inv": args.augment_include_inv,
        "augment-refine-ret": args.augment_refine_ret,
        "lambda-struct": args.lambda_struct,
    }
    for kind, prefix in WORKLOAD_PREFIXES:
        for metric in METRICS:
            augmented_mean = augmented.get(kind, {}).get(metric)
            field = f"{prefix}_{metric}_mean"
            values[field] = augmented_mean
            values[f"diff_{field}"] = (
                None if augmented_mean is None else round(baseline[kind][metric] - augmented_mean, 10)
            )
    return values


def read_existing_results(worksheet):
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    header_index = {str(header): index for index, header in enumerate(headers) if header is not None}
    results = []
    for row in rows:
        result = {
            header: row[header_index[header]] if header in header_index else None
            for header in HEADERS
        }
        if result["test_db"] not in (None, ""):
            results.append(result)
    return results


def normalized(value: object) -> str:
    return str(value if value is not None else "").strip().casefold()


def same_configuration(left: Dict[str, object], right: Dict[str, object]) -> bool:
    key_headers = ("test_db", "cardinality_type", "epochs", *CONFIG_HEADERS)
    for header in key_headers:
        if header in {"epochs", "augment-coarse-layers", "lambda-struct"}:
            if as_float(left.get(header)) != as_float(right.get(header)):
                return False
        elif header == "test-augment":
            # Rows created before TEST_AUGMENT existed always tested with augmentation enabled.
            left_value = "True" if left.get(header) in (None, "") else left.get(header)
            if normalized(left_value) != normalized(right.get(header)):
                return False
        elif normalized(left.get(header)) != normalized(right.get(header)):
            return False
    return True


def upsert_result(path: str, values: Dict[str, object]) -> None:
    output_dir = os.path.dirname(path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    if os.path.exists(path):
        workbook = load_workbook(path)
        worksheet = workbook["Augmented"] if "Augmented" in workbook.sheetnames else workbook.active
        worksheet.title = "Augmented"
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Augmented"

    results = read_existing_results(worksheet)
    target_index = next(
        (index for index, result in enumerate(results) if same_configuration(result, values)),
        None,
    )
    if target_index is None:
        results.append(values)
    else:
        results[target_index] = values

    worksheet.delete_rows(1, worksheet.max_row)
    worksheet.append(HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for result in results:
        worksheet.append(["" if result.get(header) is None else result.get(header) for header in HEADERS])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column, header in enumerate(HEADERS, start=1):
        width = max(len(header) + 2, 14 if header != "time_stamp" else 18)
        worksheet.column_dimensions[worksheet.cell(row=1, column=column).column_letter].width = width
    workbook.save(path)
    workbook.close()


def main() -> int:
    args = parse_args()
    baseline = read_baseline(Path(args.baseline_xlsx), args.test_db, args.cardinality_type)
    if baseline is None:
        print(
            "Augmented result not updated: no complete baseline row for "
            f"({args.test_db}, {args.cardinality_type}) in {args.baseline_xlsx}"
        )
        return 1

    rows, run_details = read_workbooks(args.summary_xlsx)
    summary = aggregate(rows, args.requested_runs, run_details)
    if summary["successful_runs"] == 0:
        print("Augmented result not updated: no successful augmented runs were found.")
        return 1

    values = build_values(args, summary, baseline)
    missing_metrics = [
        f"{prefix}_{metric}"
        for _, prefix in WORKLOAD_PREFIXES
        for metric in METRICS
        if values[f"{prefix}_{metric}_mean"] is None
    ]
    if missing_metrics:
        print(
            "Augmented result not updated: missing augmented workload mean(s): "
            + ", ".join(missing_metrics)
        )
        return 1
    upsert_result(args.output_xlsx, values)
    print(
        f"Augmented result updated: {args.output_xlsx} "
        f"({args.test_db}, {args.cardinality_type}, {args.augment_pooling})"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
