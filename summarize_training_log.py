#!/usr/bin/env python3
"""Print a compact summary for Graceful training logs."""

from __future__ import annotations

import argparse
import ast
import datetime as dt
import os
import re
from typing import Dict, Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


SUMMARY_START = "=== RUN SUMMARY ==="
SUMMARY_END = "=== END RUN SUMMARY ==="

FLOAT_PATTERN = r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("log_path", help="Path to the log file to summarize")
    parser.add_argument("--exit-code", type=int, default=0, help="Exit code of the training command")
    parser.add_argument("--test-db", default=None, help="Fallback test database name")
    parser.add_argument("--cardinality-type", default=None, help="Fallback cardinality type")
    parser.add_argument("--graceful-dir", default=None, help="Repository directory used by the run")
    parser.add_argument("--group-timestamp", default=None, help="Timestamp shared by repeated runs")
    parser.add_argument("--requested-runs", type=int, default=1, help="Number of requested runs")
    parser.add_argument("--run-index", type=int, default=1, help="One-based run index")
    parser.add_argument(
        "--run-variable",
        action="append",
        default=[],
        metavar="NAME=VALUE",
        help="Additional run variable to store in the Configuration sheet",
    )
    parser.add_argument("--xlsx-path", default=None, help="Optional XLSX path for the parsed run summary")
    return parser.parse_args()


def read_log(path: str) -> List[str]:
    with open(path, "r", encoding="utf-8", errors="replace") as fh:
        text = fh.read()
    text = text.split(SUMMARY_START, 1)[0]
    return text.splitlines()


def parse_dict_line(line: str, prefix: str) -> Dict[str, object]:
    if not line.startswith(prefix):
        return {}
    try:
        value = ast.literal_eval(line[len(prefix) :].strip())
    except (SyntaxError, ValueError):
        return {}
    return value if isinstance(value, dict) else {}


def fmt(value: object, digits: int = 4) -> str:
    if value is None:
        return "-"
    if isinstance(value, float):
        return f"{value:.{digits}f}"
    return str(value)


def xlsx_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def shorten(path: Optional[str]) -> str:
    if not path:
        return "-"
    return path.replace(os.path.expanduser("~"), "~")


def unique_preserve_order(values: Iterable[str]) -> List[str]:
    seen = set()
    result = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def collect_summary(lines: List[str], args: argparse.Namespace) -> Dict[str, object]:
    config: Dict[str, object] = {}
    hyperparams: Dict[str, object] = {}
    plan_counts: List[str] = []
    checkpoints: List[str] = []
    epochs: List[int] = []
    best_events: List[Dict[str, object]] = []
    final_training: Dict[str, object] = {}
    workload_metrics: List[Dict[str, object]] = []
    warning_lines: List[str] = []
    status_lines: List[str] = []

    current_epoch: Optional[int] = None
    current_epoch_metrics: Optional[Dict[str, object]] = None
    current_workload: Optional[Dict[str, object]] = None
    last_metric_name: Optional[str] = None
    last_metric_value: Optional[float] = None

    start_line = None
    done_line = None
    test_db = args.test_db
    card_type = args.cardinality_type
    model_config = None
    cuda_device = "-"
    test_all_cardinalities = "-"
    seed = "-"
    repeat_run = "-"

    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            continue

        if line.startswith("===") and "Starting" in line:
            start_line = line
        elif line.startswith("===") and "DONE" in line:
            done_line = line

        if line.startswith("Test DB:"):
            test_db = line.split(":", 1)[1].strip()
        elif line.startswith("CUDA device:"):
            cuda_device = line.split(":", 1)[1].strip()
        elif line.startswith("Test all cardinalities:"):
            test_all_cardinalities = line.split(":", 1)[1].strip()
        elif line.startswith("Seed:"):
            seed = line.split(":", 1)[1].strip()
        elif line.startswith("Repeat run:"):
            repeat_run = line.split(":", 1)[1].strip()
        elif line.startswith("Card type:"):
            card_type = line.split(":", 1)[1].strip()
        elif line.startswith("Model config:"):
            model_config = line.split(":", 1)[1].strip()

        parsed_config = parse_dict_line(line, "Running with config:")
        if parsed_config:
            config = parsed_config
        parsed_hyperparams = parse_dict_line(line, "Hyperparams:")
        if parsed_hyperparams:
            hyperparams = parsed_hyperparams

        match = re.search(r"No of Plans:\s+(\d+).*min runtime discards:\s+(\d+)", line)
        if match:
            plan_counts.append(f"{match.group(1)} plans ({match.group(2)} discards)")

        match = re.match(r"Epoch\s+(\d+)\b", line)
        if match:
            current_epoch = int(match.group(1))
            epochs.append(current_epoch)
            current_epoch_metrics = {"epoch": current_epoch}
            current_workload = None
            continue

        match = re.match(rf"val_loss epoch (\d+):\s*({FLOAT_PATTERN})", line)
        if match:
            current_epoch = int(match.group(1))
            if not epochs or epochs[-1] != current_epoch:
                epochs.append(current_epoch)
            current_epoch_metrics = {"epoch": current_epoch, "loss": float(match.group(2))}
            final_training = dict(current_epoch_metrics)
            current_workload = None
            continue

        match = re.match(r"Starting validation for\s+(.+)", line)
        if match:
            if current_workload:
                workload_metrics.append(current_workload)
            current_workload = {"workload": match.group(1).strip()}
            current_epoch_metrics = None
            continue

        match = re.match(rf"(test_\S+)_loss epoch (\d+):\s*({FLOAT_PATTERN})", line)
        if match:
            if current_workload is None:
                current_workload = {"workload": match.group(1).replace("test_", "", 1)}
            current_workload["epoch"] = int(match.group(2))
            current_workload["loss"] = float(match.group(3))
            continue

        match = re.match(rf"val_(mse|mape|median_q_error_50|median_q_error_95|median_q_error_99|procentual_error):\s*({FLOAT_PATTERN})", line)
        if match:
            metric_name = match.group(1)
            metric_value = float(match.group(2))
            last_metric_name = metric_name
            last_metric_value = metric_value
            if current_workload is not None:
                current_workload[metric_name] = metric_value
            elif current_epoch_metrics is not None:
                current_epoch_metrics[metric_name] = metric_value
                final_training = dict(current_epoch_metrics)
            continue

        if line.startswith("New best model for"):
            metric = line.split("for", 1)[1].strip()
            best_events.append(
                {
                    "epoch": current_epoch,
                    "metric": metric,
                    "value": last_metric_value if last_metric_name in metric else None,
                }
            )
            continue

        match = re.search(r"Saved checkpoint to\s+(.+?)\s+in\s+", line)
        if match:
            checkpoints.append(match.group(1))
            continue

        if "Early stopping" in line or "Training finished" in line or "Reloading best model" in line:
            status_lines.append(line)
            continue

        match = re.search(rf"qError\s*=\s*({FLOAT_PATTERN})", line)
        if match:
            if current_workload is not None:
                current_workload["worst_q_error"] = float(match.group(1))
            current_epoch_metrics = None
            continue

        lower = line.lower()
        if (
            "traceback" in lower
            or re.search(r"\berror\b", lower)
            or "exception" in lower
            or "command not found" in lower
            or "no valid checkpoint" in lower
            or "must be real number" in lower
        ):
            warning_lines.append(line)

    if current_workload:
        workload_metrics.append(current_workload)

    if config:
        test_db = str(config.get("test_against", test_db or "-"))
        card_type = str(config.get("card_type", card_type or "-"))
        model_config = str(config.get("model_config", model_config or "-"))
    if hyperparams:
        model_config = str(hyperparams.get("model_config", model_config or "-"))

    return {
        "start_line": start_line,
        "done_line": done_line,
        "test_db": test_db or "-",
        "cuda_device": cuda_device or "-",
        "test_all_cardinalities": test_all_cardinalities or "-",
        "seed": seed or "-",
        "repeat_run": repeat_run or "-",
        "card_type": card_type or "-",
        "model_config": model_config or "-",
        "data_keyword": config.get("data_keyword", hyperparams.get("data_keyword", "-")),
        "max_runtime": config.get("max_runtime", hyperparams.get("max_runtime", "-")),
        "include_pullup_data": config.get("include_pullup_data", hyperparams.get("include_pullup_data", "-")),
        "include_pushdown_data": config.get("include_pushdown_data", hyperparams.get("include_pushdown_data", "-")),
        "augment": config.get("augment", hyperparams.get("augment", "-")),
        "test_augment": config.get("test_augment", hyperparams.get("test_augment", "-")),
        "augment_pooling": config.get("augment_pooling", hyperparams.get("augment_pooling", "-")),
        "augment_refinement": config.get("augment_refinement", hyperparams.get("augment_refinement", "-")),
        "augment_coarse_layers": config.get("augment_coarse_layers", hyperparams.get("augment_coarse_layers", "-")),
        "augment_include_inv": config.get("augment_include_inv", hyperparams.get("augment_include_inv", "-")),
        "augment_refine_ret": config.get("augment_refine_ret", hyperparams.get("augment_refine_ret", "-")),
        "lambda_struct": config.get("lambda_struct", hyperparams.get("lambda_struct", "-")),
        "plan_counts": unique_preserve_order(plan_counts),
        "epochs": epochs,
        "best_events": best_events,
        "final_training": final_training,
        "workload_metrics": workload_metrics,
        "checkpoints": checkpoints,
        "status_lines": unique_preserve_order(status_lines),
        "warning_lines": unique_preserve_order(warning_lines),
        "config": config,
        "hyperparams": hyperparams,
    }


def print_summary(summary: Dict[str, object], args: argparse.Namespace) -> None:
    generated_at = dt.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")
    run_status = "SUCCESS" if args.exit_code == 0 else f"FAILED (exit code {args.exit_code})"

    print()
    print(SUMMARY_START)
    print(f"Generated at: {generated_at}")
    print(f"Run status: {run_status}")
    print(f"Log file: {shorten(args.log_path)}")
    print(f"Work dir: {shorten(args.graceful_dir)}")
    print(f"CUDA device: {summary['cuda_device']}")
    print(f"Test DB: {summary['test_db']}")
    print(f"Cardinality type: {summary['card_type']}")
    print(f"Test all cardinalities: {summary['test_all_cardinalities']}")
    print(f"Seed: {summary['seed']}")
    print(f"Repeat run: {summary['repeat_run']}")
    print(f"Model config: {summary['model_config']}")
    print(
        "Protocol: "
        f"data_keyword={summary['data_keyword']}, "
        f"max_runtime={summary['max_runtime']}, "
        f"include_pullup_data={summary['include_pullup_data']}, "
        f"include_pushdown_data={summary['include_pushdown_data']}"
    )
    #? Include augmentation settings in the end-of-log summary when they are present.
    print(
        "Augmentation: "
        f"augment={summary['augment']}, "
        f"test_augment={summary['test_augment']}, "
        f"pooling={summary['augment_pooling']}, "
        f"refinement={summary['augment_refinement']}, "
        f"coarse_layers={summary['augment_coarse_layers']}, "
        f"include_inv={summary['augment_include_inv']}, "
        f"refine_ret={summary['augment_refine_ret']}, "
        f"lambda_struct={summary['lambda_struct']}"
    )

    if summary["start_line"]:
        print(f"Started: {summary['start_line']}")
    if summary["done_line"]:
        print(f"Finished: {summary['done_line']}")

    plan_counts = summary["plan_counts"]
    if plan_counts:
        print(f"Plan counts seen: {', '.join(plan_counts[:6])}")

    epochs = summary["epochs"]
    if epochs:
        print(f"Epochs run: {len(set(epochs))} ({min(epochs)} to {max(epochs)})")

    final_training = summary["final_training"]
    if final_training:
        print(
            "Last validation metrics: "
            f"epoch={fmt(final_training.get('epoch'))}, "
            f"loss={fmt(final_training.get('loss'))}, "
            f"mse={fmt(final_training.get('mse'))}, "
            f"mape={fmt(final_training.get('mape'))}, "
            f"q50={fmt(final_training.get('median_q_error_50'))}, "
            f"q95={fmt(final_training.get('median_q_error_95'))}, "
            f"q99={fmt(final_training.get('median_q_error_99'))}"
        )

    best_events = summary["best_events"]
    if best_events:
        last_best = best_events[-1]
        print(
            "Last early-stopping improvement: "
            f"{last_best['metric']}={fmt(last_best.get('value'))} at epoch {fmt(last_best.get('epoch'))}"
        )

    checkpoints = summary["checkpoints"]
    if checkpoints:
        print(f"Last checkpoint: {shorten(checkpoints[-1])}")

    status_lines = summary["status_lines"]
    if status_lines:
        print("Status:")
        for line in status_lines[-4:]:
            print(f"  - {line}")

    workloads = summary["workload_metrics"]
    if workloads:
        print()
        print("Final workload metrics:")
        print("  workload                  loss      mse       mape      q50       q95       q99       worst_q")
        for item in workloads:
            print(
                "  "
                f"{str(item.get('workload', '-'))[:24]:24} "
                f"{fmt(item.get('loss')):>8} "
                f"{fmt(item.get('mse')):>8} "
                f"{fmt(item.get('mape')):>8} "
                f"{fmt(item.get('median_q_error_50')):>8} "
                f"{fmt(item.get('median_q_error_95')):>8} "
                f"{fmt(item.get('median_q_error_99')):>8} "
                f"{fmt(item.get('worst_q_error')):>8}"
            )

        for metric, label in [
            ("median_q_error_50", "Best q50"),
            ("median_q_error_95", "Best q95"),
            ("loss", "Lowest loss"),
            ("worst_q_error", "Lowest worst qError"),
        ]:
            with_metric = [item for item in workloads if item.get(metric) is not None]
            if with_metric:
                best = min(with_metric, key=lambda item: item[metric])
                print(f"{label}: {best['workload']} ({fmt(best[metric])})")

    warning_lines = summary["warning_lines"]
    if warning_lines:
        print()
        print(f"Warnings/errors noticed: {len(warning_lines)}")
        for line in warning_lines[-6:]:
            print(f"  - {line[:220]}")

    print(SUMMARY_END)


def best_workload(workloads: List[Dict[str, object]], metric: str) -> Optional[Dict[str, object]]:
    with_metric = [item for item in workloads if item.get(metric) is not None]
    if not with_metric:
        return None
    return min(with_metric, key=lambda item: item[metric])


def format_worksheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for column_cells in worksheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 50)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def parse_run_variables(items: Iterable[str]) -> Dict[str, str]:
    variables = {}
    for item in items:
        name, separator, value = item.partition("=")
        if not separator or not name:
            raise ValueError(f"Invalid --run-variable {item!r}; expected NAME=VALUE")
        variables[name] = value
    return variables


def write_xlsx_summary(summary: Dict[str, object], args: argparse.Namespace) -> None:
    if not args.xlsx_path:
        return

    workloads = summary["workload_metrics"] or [{}]
    best_q50 = best_workload(summary["workload_metrics"], "median_q_error_50")
    best_q95 = best_workload(summary["workload_metrics"], "median_q_error_95")
    lowest_loss = best_workload(summary["workload_metrics"], "loss")
    lowest_worst_q = best_workload(summary["workload_metrics"], "worst_q_error")
    final_training = summary["final_training"]
    checkpoints = summary["checkpoints"]

    fieldnames = [
        "run_status",
        "group_timestamp",
        "requested_runs",
        "run_index",
        "log_file",
        "work_dir",
        "cuda_device",
        "test_all_cardinalities",
        "seed",
        "repeat_run",
        "start_line",
        "done_line",
        "test_db",
        "cardinality_type",
        "model_config",
        "data_keyword",
        "max_runtime",
        "include_pullup_data",
        "include_pushdown_data",
        "augment",
        "test_augment",
        "augment_pooling",
        "augment_refinement",
        "augment_coarse_layers",
        "augment_include_inv",
        "augment_refine_ret",
        "lambda_struct",
        "epochs_run",
        "epoch_min",
        "epoch_max",
        "last_val_epoch",
        "last_val_loss",
        "last_val_mse",
        "last_val_mape",
        "last_val_q50",
        "last_val_q95",
        "last_val_q99",
        "last_checkpoint",
        "best_q50_workload",
        "best_q50",
        "best_q95_workload",
        "best_q95",
        "lowest_loss_workload",
        "lowest_loss",
        "lowest_worst_q_workload",
        "lowest_worst_q",
        "warning_count",
        "workload",
        "workload_epoch",
        "workload_loss",
        "workload_mse",
        "workload_mape",
        "workload_q50",
        "workload_q95",
        "workload_q99",
        "workload_worst_q",
    ]

    epochs = summary["epochs"]
    base_row = {
        "run_status": "SUCCESS" if args.exit_code == 0 else f"FAILED ({args.exit_code})",
        "group_timestamp": args.group_timestamp,
        "requested_runs": args.requested_runs,
        "run_index": args.run_index,
        "log_file": args.log_path,
        "work_dir": args.graceful_dir,
        "cuda_device": summary["cuda_device"],
        "test_all_cardinalities": summary["test_all_cardinalities"],
        "seed": summary["seed"],
        "repeat_run": summary["repeat_run"],
        "start_line": summary["start_line"],
        "done_line": summary["done_line"],
        "test_db": summary["test_db"],
        "cardinality_type": summary["card_type"],
        "model_config": summary["model_config"],
        "data_keyword": summary["data_keyword"],
        "max_runtime": summary["max_runtime"],
        "include_pullup_data": summary["include_pullup_data"],
        "include_pushdown_data": summary["include_pushdown_data"],
        "augment": summary["augment"],
        "test_augment": summary["test_augment"],
        "augment_pooling": summary["augment_pooling"],
        "augment_refinement": summary["augment_refinement"],
        "augment_coarse_layers": summary["augment_coarse_layers"],
        "augment_include_inv": summary["augment_include_inv"],
        "augment_refine_ret": summary["augment_refine_ret"],
        "lambda_struct": summary["lambda_struct"],
        "epochs_run": len(set(epochs)) if epochs else "",
        "epoch_min": min(epochs) if epochs else "",
        "epoch_max": max(epochs) if epochs else "",
        "last_val_epoch": final_training.get("epoch"),
        "last_val_loss": final_training.get("loss"),
        "last_val_mse": final_training.get("mse"),
        "last_val_mape": final_training.get("mape"),
        "last_val_q50": final_training.get("median_q_error_50"),
        "last_val_q95": final_training.get("median_q_error_95"),
        "last_val_q99": final_training.get("median_q_error_99"),
        "last_checkpoint": checkpoints[-1] if checkpoints else "",
        "best_q50_workload": best_q50.get("workload") if best_q50 else "",
        "best_q50": best_q50.get("median_q_error_50") if best_q50 else "",
        "best_q95_workload": best_q95.get("workload") if best_q95 else "",
        "best_q95": best_q95.get("median_q_error_95") if best_q95 else "",
        "lowest_loss_workload": lowest_loss.get("workload") if lowest_loss else "",
        "lowest_loss": lowest_loss.get("loss") if lowest_loss else "",
        "lowest_worst_q_workload": lowest_worst_q.get("workload") if lowest_worst_q else "",
        "lowest_worst_q": lowest_worst_q.get("worst_q_error") if lowest_worst_q else "",
        "warning_count": len(summary["warning_lines"]),
    }

    output_dir = os.path.dirname(args.xlsx_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    workbook = Workbook()
    results_sheet = workbook.active
    results_sheet.title = "Results"
    results_sheet.append(fieldnames)
    for workload in workloads:
        row = dict(base_row)
        row.update(
            {
                "workload": workload.get("workload"),
                "workload_epoch": workload.get("epoch"),
                "workload_loss": workload.get("loss"),
                "workload_mse": workload.get("mse"),
                "workload_mape": workload.get("mape"),
                "workload_q50": workload.get("median_q_error_50"),
                "workload_q95": workload.get("median_q_error_95"),
                "workload_q99": workload.get("median_q_error_99"),
                "workload_worst_q": workload.get("worst_q_error"),
            }
        )
        results_sheet.append([xlsx_value(row.get(key)) for key in fieldnames])
    format_worksheet(results_sheet)

    configuration_sheet = workbook.create_sheet("Configuration")
    configuration_sheet.append(["source", "variable", "value"])
    metadata = {
        "group_timestamp": args.group_timestamp,
        "requested_runs": args.requested_runs,
        "run_index": args.run_index,
        **base_row,
    }
    for variable, value in metadata.items():
        configuration_sheet.append(["run", variable, xlsx_value(value)])
    for variable, value in parse_run_variables(args.run_variable).items():
        configuration_sheet.append(["run_variable", variable, xlsx_value(value)])
    for source in ("config", "hyperparams"):
        for variable, value in sorted(summary[source].items()):
            configuration_sheet.append([source, variable, xlsx_value(value)])
    format_worksheet(configuration_sheet)

    workbook.save(args.xlsx_path)
    print(f"XLSX summary: {args.xlsx_path}")


def main() -> int:
    args = parse_args()
    lines = read_log(args.log_path)
    summary = collect_summary(lines, args)
    print_summary(summary, args)
    write_xlsx_summary(summary, args)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
