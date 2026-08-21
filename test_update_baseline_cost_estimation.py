import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from update_baseline_cost_estimation import HEADERS, upsert_result


def result(epochs):
    return {
        "test_db": "accidents",
        "cardinality_type": "est",
        "pull_up_q50_mean": 1.2,
        "pull_up_q95_mean": 5.0,
        "pull_up_q99_mean": 10.0,
        "push_down_q50_mean": 1.4,
        "push_down_q95_mean": 7.0,
        "push_down_q99_mean": 20.0,
        "time_stamp": "20260821_120000",
        "epochs": epochs,
    }


class UpdateBaselineCostEstimationTest(unittest.TestCase):
    def test_epochs_is_informational_and_does_not_change_upsert_key(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "baseline.xlsx"
            upsert_result(str(path), result(50))
            upsert_result(str(path), result(100))

            workbook = load_workbook(path, read_only=True, data_only=True)
            rows = list(workbook["Baseline"].iter_rows(values_only=True))
            workbook.close()

            self.assertEqual(list(rows[0]), HEADERS)
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[1][HEADERS.index("epochs")], 100)



if __name__ == "__main__":
    unittest.main()
