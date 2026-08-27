import argparse
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from update_augmented_cost_estimation import (
    HEADERS,
    build_values,
    read_baseline,
    same_configuration,
    upsert_result,
)


def arguments(**overrides):
    values = {
        "test_db": "accidents",
        "cardinality_type": "est",
        "time_stamp": "20260820_120000",
        "epochs": 100,
        "test_augment": "True",
        "augment_pooling": "hybrid",
        "augment_refinement": "gated_residual",
        "augment_coarse_layers": 1,
        "augment_include_inv": "False",
        "augment_refine_ret": "False",
        "lambda_struct": 0.0,
    }
    values.update(overrides)
    return argparse.Namespace(**values)


class UpdateAugmentedCostEstimationTest(unittest.TestCase):
    def test_legacy_config_defaults_test_augment_to_true(self):
        current = {
            "test_db": "accidents",
            "cardinality_type": "est",
            "epochs": 100,
            "test-augment": "True",
            "augment-pooling": "hybrid",
            "augment-refinement": "gated_residual",
            "augment-coarse-layers": 1,
            "augment-include-inv": "False",
            "augment-refine-ret": "False",
            "lambda-struct": 0.0,
        }
        legacy = {**current, "test-augment": None}

        self.assertTrue(same_configuration(legacy, current))
        self.assertFalse(same_configuration(legacy, {**current, "test-augment": "False"}))

    def test_read_baseline_matches_database_and_cardinality(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "baseline.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Baseline"
            worksheet.append([
                "test_db", "cardinality_type",
                "pull_up_q50_mean", "pull_up_q95_mean", "pull_up_q99_mean",
                "push_down_q50_mean", "push_down_q95_mean", "push_down_q99_mean",
                "time_stamp",
            ])
            worksheet.append(["accidents", "act", 20.0, 30.0, 40.0, 50.0, 60.0, 70.0, "old"])
            worksheet.append(["accidents", "est", 4.0, 7.0, 10.0, 5.0, 8.0, 11.0, "new"])
            workbook.save(path)
            workbook.close()

            actual = read_baseline(path, "ACCIDENTS", "EST")

            self.assertEqual(actual["pullup"], {"q50": 4.0, "q95": 7.0, "q99": 10.0})
            self.assertEqual(actual["pushdown"], {"q50": 5.0, "q95": 8.0, "q99": 11.0})

    def test_build_values_computes_baseline_minus_augmented(self):
        summary = {
            "workloads": {
                "workload_pullup_est": {
                    "q50": (2.5, 0.1, 2), "q95": (5.0, 0.2, 2), "q99": (8.5, 0.3, 2),
                },
                "workload_pushdown_est": {
                    "q50": (3.0, 0.1, 2), "q95": (6.0, 0.2, 2), "q99": (9.0, 0.3, 2),
                },
            }
        }

        actual = build_values(
            arguments(),
            summary,
            {
                "pullup": {"q50": 4.0, "q95": 7.0, "q99": 10.0},
                "pushdown": {"q50": 5.0, "q95": 8.0, "q99": 11.0},
            },
        )

        self.assertEqual(actual["diff_pull_up_q50_mean"], 1.5)
        self.assertEqual(actual["diff_pull_up_q95_mean"], 2.0)
        self.assertEqual(actual["diff_pull_up_q99_mean"], 1.5)
        self.assertEqual(actual["diff_push_down_q50_mean"], 2.0)

    def test_upsert_keeps_distinct_configs_and_replaces_matching_config(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "augmented.xlsx"
            summary = {
                "workloads": {
                    "workload_pullup_est": {
                        "q50": (2.5, 0.0, 1), "q95": (5.0, 0.0, 1), "q99": (8.5, 0.0, 1),
                    },
                    "workload_pushdown_est": {
                        "q50": (3.0, 0.0, 1), "q95": (6.0, 0.0, 1), "q99": (9.0, 0.0, 1),
                    },
                }
            }
            baseline = {
                "pullup": {"q50": 4.0, "q95": 7.0, "q99": 10.0},
                "pushdown": {"q50": 5.0, "q95": 8.0, "q99": 11.0},
            }
            first = build_values(arguments(), summary, baseline)
            replacement = build_values(arguments(time_stamp="20260820_130000"), summary, baseline)
            distinct_epochs = build_values(arguments(epochs=200), summary, baseline)
            distinct = build_values(arguments(augment_pooling="max"), summary, baseline)
            no_test_augment = build_values(arguments(test_augment="False"), summary, baseline)

            upsert_result(str(path), first)
            upsert_result(str(path), replacement)
            upsert_result(str(path), distinct_epochs)
            upsert_result(str(path), distinct)
            upsert_result(str(path), no_test_augment)

            workbook = load_workbook(path, read_only=True, data_only=True)
            worksheet = workbook["Augmented"]
            rows = list(worksheet.iter_rows(values_only=True))
            workbook.close()
            self.assertEqual(list(rows[0]), HEADERS)
            self.assertEqual(len(rows), 5)
            self.assertEqual(rows[1][HEADERS.index("time_stamp")], "20260820_130000")
            self.assertEqual(rows[1][HEADERS.index("epochs")], 100)
            self.assertEqual(rows[2][HEADERS.index("epochs")], 200)
            self.assertEqual(rows[3][HEADERS.index("augment-pooling")], "max")
            self.assertEqual(rows[4][HEADERS.index("test-augment")], "False")


if __name__ == "__main__":
    unittest.main()
